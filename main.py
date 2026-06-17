import os
import sys
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QFrame, QPushButton, QLabel, QLineEdit, QTextEdit, QTextBrowser,
    QFileDialog, QCheckBox, QScrollArea, QSplitter,
    QMessageBox, QTabWidget
)
from PySide6.QtCore import Qt, Signal, Slot

import config
from document_parser import extract_text
from groq_client import GroqChatThread
from exporter import export_to_word, export_to_pdf, export_to_excel


# ═══════════════════════════════════════════════════════════════════════════════
#  Shared helper widgets
# ═══════════════════════════════════════════════════════════════════════════════

class ChatTextEdit(QTextEdit):
    """Enter to send, Shift+Enter for newline."""
    enter_pressed = Signal()

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            if event.modifiers() & Qt.ShiftModifier:
                super().keyPressEvent(event)
            else:
                self.enter_pressed.emit()
        else:
            super().keyPressEvent(event)


class FileItemWidget(QFrame):
    """Sidebar card for one uploaded document."""
    state_changed = Signal(bool, str)
    deleted       = Signal(str)

    def __init__(self, file_path: str):
        super().__init__()
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        try:
            b = os.path.getsize(file_path)
            self.size_str = f"{b/1024:.1f} KB" if b < 1024 ** 2 else f"{b/1024**2:.1f} MB"
        except Exception:
            self.size_str = "Unknown"

        self.setObjectName("FileItemWidget")
        self.setFrameShape(QFrame.StyledPanel)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(6, 6, 6, 6)
        lay.setSpacing(6)

        self.checkbox = QCheckBox()
        self.checkbox.setChecked(True)
        self.checkbox.stateChanged.connect(
            lambda s: self.state_changed.emit(s == Qt.Checked or s == 2, self.file_path)
        )
        lay.addWidget(self.checkbox)

        details = QVBoxLayout()
        details.setSpacing(2)
        name_lbl = QLabel(self.file_name)
        name_lbl.setStyleSheet("font-weight:bold;color:#eceff4;font-size:11px;")
        name_lbl.setWordWrap(True)
        size_lbl = QLabel(self.size_str)
        size_lbl.setStyleSheet("color:#989fb5;font-size:9px;")
        details.addWidget(name_lbl)
        details.addWidget(size_lbl)
        lay.addLayout(details, 1)

        del_btn = QPushButton("✕")
        del_btn.setFixedSize(22, 22)
        del_btn.setObjectName("DeleteFileButton")
        del_btn.setToolTip("Remove from session")
        del_btn.clicked.connect(lambda: self.deleted.emit(self.file_path))
        lay.addWidget(del_btn)


class UserMessageWidget(QWidget):
    """Right-aligned user bubble."""
    def __init__(self, text: str):
        super().__init__()
        lay = QHBoxLayout(self)
        lay.setContentsMargins(10, 5, 10, 5)
        lay.addStretch()
        bubble = QFrame()
        bubble.setObjectName("UserBubble")
        bubble.setFrameShape(QFrame.StyledPanel)
        bl = QVBoxLayout(bubble)
        bl.setContentsMargins(14, 10, 14, 10)
        lbl = QLabel(text)
        lbl.setWordWrap(True)
        lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
        lbl.setStyleSheet("color:#eceff4;font-size:13px;")
        bl.addWidget(lbl)
        lay.addWidget(bubble)
        bubble.setMaximumWidth(500)


class AIMessageWidget(QWidget):
    """Left-aligned AI bubble with Markdown rendering and per-response export."""
    export_clicked = Signal(str, str, str)

    def __init__(self, original_prompt: str):
        super().__init__()
        self.original_prompt = original_prompt
        self.md_content = ""

        lay = QHBoxLayout(self)
        lay.setContentsMargins(10, 5, 10, 5)

        self.bubble = QFrame()
        self.bubble.setObjectName("AIBubble")
        self.bubble.setFrameShape(QFrame.StyledPanel)
        bl = QVBoxLayout(self.bubble)
        bl.setContentsMargins(14, 10, 14, 10)
        bl.setSpacing(8)

        self.text_browser = QTextBrowser()
        self.text_browser.setObjectName("AITextBrowser")
        self.text_browser.setOpenExternalLinks(True)
        self.text_browser.setFrameStyle(QFrame.NoFrame)
        self.text_browser.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.text_browser.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.text_browser.textChanged.connect(self._adjust_height)
        self.text_browser.setStyleSheet(
            "QTextBrowser{background:transparent;border:none;font-size:13px;color:#eceff4;}"
        )
        self.text_browser.document().setDefaultStyleSheet(
            "h1{color:#89b4fa;font-size:16px;margin-top:10px;margin-bottom:5px;font-weight:bold;}"
            "h2{color:#89b4fa;font-size:14px;margin-top:8px;margin-bottom:4px;font-weight:bold;}"
            "h3{color:#cba6f7;font-size:12px;margin-top:6px;margin-bottom:3px;font-weight:bold;}"
            "p{line-height:1.4;margin-bottom:6px;}"
            "ul,ol{margin-left:15px;margin-bottom:6px;}"
            "li{margin-bottom:3px;}"
            "code{font-family:'Courier New',monospace;background:#2e3440;padding:2px 4px;"
            "     color:#f38ba8;border-radius:3px;}"
            "pre{background:#1e1e2e;border:1px solid #313244;padding:8px;"
            "    font-family:'Courier New',monospace;border-radius:4px;}"
            "table{border-collapse:collapse;margin-top:10px;margin-bottom:10px;}"
            "th,td{border:1px solid #45475a;padding:5px 10px;}"
            "th{background:#313244;font-weight:bold;color:#89b4fa;}"
        )
        bl.addWidget(self.text_browser)

        # Export action bar
        self.actions_widget = QWidget()
        al = QHBoxLayout(self.actions_widget)
        al.setContentsMargins(0, 4, 0, 0)
        al.setSpacing(8)
        exp_lbl = QLabel("Export this response:")
        exp_lbl.setStyleSheet("color:#6c7086;font-size:10px;")
        al.addWidget(exp_lbl)
        for label, fmt in [("📄 Word", "word"), ("🖨️ PDF", "pdf"), ("📊 Excel", "excel")]:
            btn = QPushButton(label)
            btn.setObjectName("BubbleExportButton")
            btn.setFixedHeight(22)
            btn.clicked.connect(
                lambda checked, f=fmt: self.export_clicked.emit(f, self.md_content, self.original_prompt)
            )
            al.addWidget(btn)
        al.addStretch()
        bl.addWidget(self.actions_widget)
        self.actions_widget.setVisible(False)

        lay.addWidget(self.bubble)
        lay.addStretch()
        self.bubble.setMaximumWidth(550)

    def append_text(self, chunk: str):
        self.md_content += chunk
        self.text_browser.setMarkdown(self.md_content)
        if not self.actions_widget.isVisible() and self.md_content.strip():
            self.actions_widget.setVisible(True)

    def set_content(self, text: str):
        self.md_content = text
        self.text_browser.setMarkdown(text)
        self.actions_widget.setVisible(True)

    def _adjust_height(self):
        self.text_browser.document().setTextWidth(self.text_browser.width())
        h = int(self.text_browser.document().size().height()) + 12
        self.text_browser.setFixedHeight(h)


# ═══════════════════════════════════════════════════════════════════════════════
#  ChatPanel — reusable self-contained chat panel
# ═══════════════════════════════════════════════════════════════════════════════

class ChatPanel(QWidget):
    """
    Full chat panel (header + scroll area + typing bar + input bar).

    Parameters
    ----------
    system_prompt       : base system message sent on every request
    context_provider    : callable() -> str | None  – called before each send
                          to inject extra context (e.g. document text)
    attach_callback     : if provided, a 📎 button is shown in the input bar
    header_right_widget : optional widget placed on the right of the header
    """

    def __init__(
        self,
        system_prompt: str,
        context_provider=None,
        attach_callback=None,
        header_right_widget=None,
    ):
        super().__init__()
        self.system_prompt      = system_prompt
        self.context_provider   = context_provider
        self.chat_history: list = []
        self.current_thread     = None
        self.active_ai_bubble   = None
        self.extra_disable_widgets: list = []  # populated by parent tabs
        self.setObjectName("ChatPanel")

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # ── Header ──────────────────────────────────────────────────────────
        header = QFrame()
        header.setObjectName("ChatHeader")
        header.setFrameShape(QFrame.StyledPanel)
        header.setFixedHeight(50)
        hl = QHBoxLayout(header)
        hl.setContentsMargins(15, 0, 15, 0)

        self.status_indicator = QLabel("● Ready")
        self.status_indicator.setObjectName("StatusIndicator")
        self.status_indicator.setStyleSheet("color:#a6e3a1;font-weight:bold;font-size:12px;")
        hl.addWidget(self.status_indicator)
        hl.addStretch()

        if header_right_widget:
            hl.addWidget(header_right_widget)

        lay.addWidget(header)

        # ── Chat scroll area ─────────────────────────────────────────────────
        self.chat_scroll = QScrollArea()
        self.chat_scroll.setWidgetResizable(True)
        self.chat_scroll.setObjectName("ChatScrollArea")
        self.chat_scroll.setFrameStyle(QFrame.NoFrame)

        self.chat_content = QWidget()
        self.chat_content.setObjectName("ChatContent")
        self.chat_list_layout = QVBoxLayout(self.chat_content)
        self.chat_list_layout.setContentsMargins(10, 10, 10, 10)
        self.chat_list_layout.setSpacing(12)
        self.chat_list_layout.addStretch()

        self.chat_scroll.setWidget(self.chat_content)
        lay.addWidget(self.chat_scroll, 1)

        # ── Typing indicator ─────────────────────────────────────────────────
        self.typing_indicator = QFrame()
        self.typing_indicator.setFixedHeight(30)
        self.typing_indicator.setObjectName("TypingIndicator")
        tl = QHBoxLayout(self.typing_indicator)
        tl.setContentsMargins(15, 0, 15, 0)
        ti_lbl = QLabel("AI is writing a response...")
        ti_lbl.setStyleSheet("color:#f9e2af;font-size:11px;font-style:italic;")
        tl.addWidget(ti_lbl)
        tl.addStretch()
        lay.addWidget(self.typing_indicator)
        self.typing_indicator.setVisible(False)

        # ── Input panel ──────────────────────────────────────────────────────
        input_panel = QFrame()
        input_panel.setObjectName("InputPanel")
        input_panel.setFrameShape(QFrame.StyledPanel)
        il = QHBoxLayout(input_panel)
        il.setContentsMargins(12, 10, 12, 12)
        il.setSpacing(8)

        if attach_callback:
            self.attach_btn = QPushButton("📎")
            self.attach_btn.setFixedSize(36, 36)
            self.attach_btn.setObjectName("AttachButton")
            self.attach_btn.setToolTip("Upload and attach a document")
            self.attach_btn.clicked.connect(attach_callback)
            il.addWidget(self.attach_btn)
        else:
            self.attach_btn = None

        self.message_input = ChatTextEdit()
        self.message_input.setPlaceholderText(
            "Type your prompt here… (Shift+Enter for new line, Enter to send)"
        )
        self.message_input.setObjectName("MessageInput")
        self.message_input.setFixedHeight(42)
        self.message_input.enter_pressed.connect(self.send_message)
        il.addWidget(self.message_input, 1)

        self.send_btn = QPushButton("Send")
        self.send_btn.setObjectName("SendButton")
        self.send_btn.setFixedSize(70, 36)
        self.send_btn.clicked.connect(self.send_message)
        il.addWidget(self.send_btn)

        lay.addWidget(input_panel)

    # ── Message building ────────────────────────────────────────────────────

    def _build_messages(self, prompt: str) -> list:
        messages = [{"role": "system", "content": self.system_prompt}]
        for role, msg in self.chat_history:
            messages.append({"role": role, "content": msg})
        if self.context_provider:
            ctx = self.context_provider()
            if ctx:
                messages.append({"role": "system", "content": ctx})
        messages.append({"role": "user", "content": prompt})
        return messages

    # ── Send / stream ────────────────────────────────────────────────────────

    @Slot()
    def send_message(self):
        prompt = self.message_input.toPlainText().strip()
        if not prompt:
            return

        self.chat_list_layout.insertWidget(
            self.chat_list_layout.count() - 1, UserMessageWidget(prompt)
        )
        self.message_input.clear()

        messages = self._build_messages(prompt)
        self.chat_history.append(("user", prompt))

        self.active_ai_bubble = AIMessageWidget(original_prompt=prompt)
        self.active_ai_bubble.export_clicked.connect(self.on_bubble_export_clicked)
        self.chat_list_layout.insertWidget(
            self.chat_list_layout.count() - 1, self.active_ai_bubble
        )

        self.scroll_to_bottom()
        self.set_controls_enabled(False)
        self.status_indicator.setText("● Generating…")
        self.status_indicator.setStyleSheet("color:#f9e2af;font-weight:bold;font-size:12px;")
        self.typing_indicator.setVisible(True)

        self.current_thread = GroqChatThread(
            config.DEFAULT_API_KEY, config.DEFAULT_MODEL,
            messages, config.DEFAULT_TEMPERATURE
        )
        self.current_thread.token_received.connect(self.on_token_received)
        self.current_thread.finished.connect(self.on_generation_finished)
        self.current_thread.error.connect(self.on_generation_error)
        self.current_thread.start()

    @Slot(str)
    def on_token_received(self, token: str):
        if self.active_ai_bubble:
            self.active_ai_bubble.append_text(token)
            self.scroll_to_bottom()

    @Slot(str)
    def on_generation_finished(self, full_text: str):
        self.chat_history.append(("assistant", full_text))
        self._cleanup_thread()
        self.scroll_to_bottom()

    @Slot(str)
    def on_generation_error(self, error_message: str):
        if self.active_ai_bubble:
            self.active_ai_bubble.set_content(f"⚠️ **Error**:\n\n`{error_message}`")
        else:
            QMessageBox.critical(self.window(), "Error", error_message)
        if self.chat_history and self.chat_history[-1][0] == "user":
            self.chat_history.pop()
        self._cleanup_thread()

    def _cleanup_thread(self):
        self.set_controls_enabled(True)
        self.typing_indicator.setVisible(False)
        self.status_indicator.setText("● Ready")
        self.status_indicator.setStyleSheet("color:#a6e3a1;font-weight:bold;font-size:12px;")
        if self.current_thread:
            self.current_thread.deleteLater()
            self.current_thread = None

    def set_controls_enabled(self, enabled: bool):
        self.send_btn.setEnabled(enabled)
        self.message_input.setEnabled(enabled)
        if self.attach_btn:
            self.attach_btn.setEnabled(enabled)
        for w in self.extra_disable_widgets:
            w.setEnabled(enabled)

    # ── Helpers ─────────────────────────────────────────────────────────────

    def scroll_to_bottom(self):
        QApplication.processEvents()
        self.chat_scroll.verticalScrollBar().setValue(
            self.chat_scroll.verticalScrollBar().maximum()
        )

    def add_system_message(self, text: str):
        bubble = AIMessageWidget(original_prompt="")
        bubble.set_content(text)
        bubble.actions_widget.setVisible(False)
        self.chat_list_layout.insertWidget(self.chat_list_layout.count() - 1, bubble)
        self.scroll_to_bottom()

    @Slot()
    def clear_chat(self):
        self.chat_history.clear()
        while self.chat_list_layout.count() > 1:
            child = self.chat_list_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        self.add_system_message("Conversation cleared. Start a new session.")

    # ── Per-bubble export ────────────────────────────────────────────────────

    @Slot(str, str, str)
    def on_bubble_export_clicked(self, format_type: str, response_text: str, original_prompt: str):
        if not response_text.strip():
            QMessageBox.warning(self.window(), "Export Warning", "Cannot export an empty response.")
            return

        if format_type == "word":
            ff, ext = "Word Document (*.docx)", ".docx"
            fn = lambda p: export_to_word(response_text, p, "AI Response")
        elif format_type == "pdf":
            ff, ext = "PDF Document (*.pdf)", ".pdf"
            fn = lambda p: export_to_pdf(response_text, p)
        elif format_type == "excel":
            ff, ext = "Excel Spreadsheet (*.xlsx)", ".xlsx"
            fn = lambda p: export_to_excel(response_text, p, q_text=original_prompt)
        else:
            return

        path, _ = QFileDialog.getSaveFileName(self.window(), "Export Response", "", ff)
        if path:
            if not path.endswith(ext):
                path += ext
            try:
                fn(path)
                QMessageBox.information(
                    self.window(), "Export Success",
                    f"Saved to:\n{os.path.basename(path)}"
                )
            except Exception as e:
                QMessageBox.critical(self.window(), "Export Failed", str(e))


# ═══════════════════════════════════════════════════════════════════════════════
#  DocuIntelligence ChatTab — document Q&A with sidebar
# ═══════════════════════════════════════════════════════════════════════════════

class DocuIntelligenceTab(QWidget):
    """
    Full document Q&A experience: left sidebar for uploading / toggling files,
    right ChatPanel that injects selected document text into every request.
    """

    def __init__(self):
        super().__init__()
        self.uploaded_files:      dict = {}
        self.active_context_files: set = set()

        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        lay.addWidget(splitter)

        # Build sidebar first so self.upload_btn exists before ChatPanel
        sidebar = self._build_sidebar()
        splitter.addWidget(sidebar)

        # Header-right: active document summary label
        self.active_docs_summary = QLabel("No active documents")
        self.active_docs_summary.setStyleSheet(
            "color:#bac2de;font-size:11px;font-style:italic;"
        )

        self.chat_panel = ChatPanel(
            system_prompt=config.DEFAULT_SYSTEM_PROMPT,
            context_provider=self._get_document_context,
            attach_callback=self.on_upload_click,
            header_right_widget=self.active_docs_summary,
        )
        # Disable upload button during generation
        self.chat_panel.extra_disable_widgets = [self.upload_btn]

        splitter.addWidget(self.chat_panel)
        splitter.setSizes([300, 700])

        self.chat_panel.add_system_message(
            "**DocuIntelligence AI** is ready. Upload a document on the left sidebar to start your analysis."
        )

    # ── Sidebar builder ─────────────────────────────────────────────────────

    def _build_sidebar(self) -> QWidget:
        sb = QWidget()
        sb.setObjectName("Sidebar")
        sb.setMinimumWidth(260)
        sb.setMaximumWidth(350)
        sl = QVBoxLayout(sb)
        sl.setContentsMargins(12, 12, 12, 12)
        sl.setSpacing(12)

        # Brand
        brand = QLabel("DocuIntelligence AI")
        brand.setObjectName("BrandLabel")
        sl.addWidget(brand)

        # Instruction card
        card = QFrame()
        card.setObjectName("InstructionCard")
        cl = QVBoxLayout(card)
        cl.setContentsMargins(12, 12, 12, 12)
        cl.setSpacing(6)
        welcome = QLabel("Welcome to DocuIntelligence AI!")
        welcome.setStyleSheet("font-weight:bold;color:#89b4fa;font-size:13px;")
        cl.addWidget(welcome)
        desc = QLabel(
            "Upload documentation files (PDF, Word, Excel, or Text) below and toggle "
            "checkboxes to include them in your query context.\n\n"
            "Answers can be exported to Word, PDF, or Excel."
        )
        desc.setWordWrap(True)
        desc.setStyleSheet("color:#a6adc8;font-size:11px;")
        cl.addWidget(desc)
        sl.addWidget(card)

        # Upload section
        sl.addWidget(self._section_lbl("DOCUMENT KNOWLEDGE BASE"))

        self.upload_btn = QPushButton("+ Upload Document")
        self.upload_btn.setObjectName("UploadButton")
        self.upload_btn.clicked.connect(self.on_upload_click)
        sl.addWidget(self.upload_btn)

        # File list scroll area
        files_scroll = QScrollArea()
        files_scroll.setWidgetResizable(True)
        files_scroll.setObjectName("FilesScrollArea")
        self.files_container = QWidget()
        self.files_container.setObjectName("FilesContainer")
        self.files_layout = QVBoxLayout(self.files_container)
        self.files_layout.setContentsMargins(4, 4, 4, 4)
        self.files_layout.setSpacing(6)
        self.files_layout.addStretch()
        files_scroll.setWidget(self.files_container)
        sl.addWidget(files_scroll, 1)

        # Clear chat
        clear_btn = QPushButton("🗑️ Clear Chat Conversation")
        clear_btn.setObjectName("SecondaryButton")
        clear_btn.clicked.connect(lambda: self.chat_panel.clear_chat())
        sl.addWidget(clear_btn)

        # Full-thread export
        sl.addWidget(self._section_lbl("EXPORT ENTIRE THREAD"))
        exp_lay = QHBoxLayout()
        for label, fmt in [("Word", "word"), ("PDF", "pdf"), ("Excel", "excel")]:
            btn = QPushButton(label)
            btn.setObjectName("BubbleExportButton")
            btn.setFixedHeight(28)
            btn.clicked.connect(lambda checked, f=fmt: self.export_full_chat(f))
            exp_lay.addWidget(btn)
        sl.addLayout(exp_lay)

        return sb

    def _section_lbl(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setObjectName("SectionTitle")
        return lbl

    # ── Document context ────────────────────────────────────────────────────

    def _get_document_context(self) -> str | None:
        docs = [
            self.uploaded_files[p]
            for p in self.active_context_files
            if p in self.uploaded_files
        ]
        if not docs:
            return None
        pieces = [f"--- DOCUMENT: {d['name']} ---\n{d['content']}" for d in docs]
        return (
            "The following text is parsed from active uploaded files in the knowledge base. "
            "Read the document context below and answer the user query using this data:\n\n"
            + "\n\n".join(pieces)
            + "\n\nAnswer the next query using the guidelines and facts from this context."
        )

    # ── Upload / file management ────────────────────────────────────────────

    @Slot()
    def on_upload_click(self):
        ff = (
            "All Supported Files (*.pdf *.docx *.xlsx *.xls *.txt *.md *.csv);;"
            "PDF Files (*.pdf);;Word Files (*.docx);;Excel Files (*.xlsx *.xls);;"
            "Text Files (*.txt *.md *.csv)"
        )
        paths, _ = QFileDialog.getOpenFileNames(self, "Upload Knowledge Base Documents", "", ff)
        if not paths:
            return

        for path in paths:
            if path in self.uploaded_files:
                continue
            self.chat_panel.status_indicator.setText("● Loading…")
            self.chat_panel.status_indicator.setStyleSheet(
                "color:#f9e2af;font-weight:bold;font-size:12px;"
            )
            QApplication.processEvents()
            try:
                content = extract_text(path)
                name = os.path.basename(path)
                widget = FileItemWidget(path)
                widget.state_changed.connect(self.on_file_state_changed)
                widget.deleted.connect(self.on_file_deleted)
                self.files_layout.insertWidget(self.files_layout.count() - 1, widget)
                self.uploaded_files[path] = {"name": name, "content": content, "widget": widget}
                self.active_context_files.add(path)
                self.chat_panel.add_system_message(
                    f"📂 **Attached**: `{name}` ({widget.size_str}) added to session context."
                )
            except Exception as e:
                QMessageBox.critical(
                    self, "Import Error",
                    f"Failed to parse '{os.path.basename(path)}':\n{e}"
                )

        self._update_docs_summary()
        self.chat_panel.status_indicator.setText("● Ready")
        self.chat_panel.status_indicator.setStyleSheet(
            "color:#a6e3a1;font-weight:bold;font-size:12px;"
        )

    @Slot(bool, str)
    def on_file_state_changed(self, is_checked: bool, file_path: str):
        if is_checked:
            self.active_context_files.add(file_path)
        else:
            self.active_context_files.discard(file_path)
        self._update_docs_summary()

    @Slot(str)
    def on_file_deleted(self, file_path: str):
        if file_path not in self.uploaded_files:
            return
        name   = self.uploaded_files[file_path]["name"]
        widget = self.uploaded_files[file_path]["widget"]
        self.files_layout.removeWidget(widget)
        widget.deleteLater()
        del self.uploaded_files[file_path]
        self.active_context_files.discard(file_path)
        self.chat_panel.add_system_message(f"🗑️ Removed `{name}` from session context.")
        self._update_docs_summary()

    def _update_docs_summary(self):
        n = len(self.active_context_files)
        if n == 0:
            self.active_docs_summary.setText("No active documents")
        elif n == 1:
            p = next(iter(self.active_context_files))
            self.active_docs_summary.setText(f"Active: {self.uploaded_files[p]['name']}")
        else:
            self.active_docs_summary.setText(f"Active: {n} documents selected")

    # ── Full-thread export ───────────────────────────────────────────────────

    def export_full_chat(self, format_type: str):
        history = self.chat_panel.chat_history
        if not history:
            QMessageBox.warning(self, "Export Warning", "Chat history is empty.")
            return

        full_md, excel_rows = [], []
        for role, msg in history:
            label = "User Query" if role == "user" else "AI Answer"
            full_md.append(f"# {label}\n\n{msg}\n")
            excel_rows.append((label, msg))
        markdown_full = "\n".join(full_md)

        if format_type == "word":
            ff, ext = "Word Document (*.docx)", ".docx"
            fn = lambda p: export_to_word(markdown_full, p, "Full Chat Transcript")
        elif format_type == "pdf":
            ff, ext = "PDF Document (*.pdf)", ".pdf"
            fn = lambda p: export_to_pdf(markdown_full, p)
        elif format_type == "excel":
            ff, ext = "Excel Spreadsheet (*.xlsx)", ".xlsx"
            def fn(path):
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                import re
                wb = Workbook()
                ws = wb.active
                ws.title = "Chat History"
                hfill = PatternFill(start_color="1E1E2E", end_color="1E1E2E", fill_type="solid")
                hfont = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                tfont = Font(name="Segoe UI", size=10)
                bfont = Font(name="Segoe UI", size=10, bold=True)
                al    = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ac    = Alignment(horizontal="center", vertical="center", wrap_text=True)
                bdr   = Border(
                    left=Side(style="thin", color="D2D2D2"),
                    right=Side(style="thin", color="D2D2D2"),
                    top=Side(style="thin", color="D2D2D2"),
                    bottom=Side(style="thin", color="D2D2D2"),
                )
                ws.column_dimensions["A"].width = 15
                ws.column_dimensions["B"].width = 85
                for col, val in [(1, "Role"), (2, "Message")]:
                    c = ws.cell(row=1, column=col, value=val)
                    c.fill, c.font, c.alignment = hfill, hfont, ac
                for i, (rl, msg) in enumerate(excel_rows):
                    r = i + 2
                    ca = ws.cell(row=r, column=1, value=rl)
                    ca.font, ca.alignment, ca.border = bfont, al, bdr
                    cb = ws.cell(row=r, column=2, value=re.sub(r"\*\*|__|\*|_|`", "", msg))
                    cb.font, cb.alignment, cb.border = tfont, al, bdr
                wb.save(path)
        else:
            return

        path, _ = QFileDialog.getSaveFileName(self, "Export Full Chat", "", ff)
        if path:
            if not path.endswith(ext):
                path += ext
            try:
                fn(path)
                QMessageBox.information(
                    self, "Export Success",
                    f"Transcript saved to:\n{os.path.basename(path)}"
                )
            except Exception as e:
                QMessageBox.critical(self, "Export Failed", str(e))


# ═══════════════════════════════════════════════════════════════════════════════
#  DocuAssistanceTab — general Q&A without document upload
# ═══════════════════════════════════════════════════════════════════════════════

class DocuAssistanceTab(QWidget):
    """
    Direct Q&A tab. Same chat UI as DocuIntelligenceTab but without the document
    sidebar. Uses a general-purpose system prompt.
    """

    def __init__(self):
        super().__init__()
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Header-right: label + clear button
        hr_widget = QWidget()
        hr_lay = QHBoxLayout(hr_widget)
        hr_lay.setContentsMargins(0, 0, 0, 0)
        hr_lay.setSpacing(12)

        mode_lbl = QLabel("Query your Docu Assistant")
        mode_lbl.setStyleSheet("color:#bac2de;font-size:11px;font-style:italic;")
        hr_lay.addWidget(mode_lbl)

        # Create the chat panel
        self.chat_panel = ChatPanel(
            system_prompt=config.KNOWLEDGE_HUB_SYSTEM_PROMPT,
            header_right_widget=hr_widget,
        )
        lay.addWidget(self.chat_panel)

        # Clear button (added to hr_widget after chat_panel exists)
        clear_btn = QPushButton("🗑️ Clear")
        clear_btn.setObjectName("SecondaryButton")
        clear_btn.setFixedHeight(26)
        clear_btn.clicked.connect(self.chat_panel.clear_chat)
        hr_lay.addWidget(clear_btn)

        self.chat_panel.add_system_message(
            "**Docu Assistant** — Ask me anything. "
            "I'll answer from my knowledge repository."
        )


# ═══════════════════════════════════════════════════════════════════════════════
#  MainWindow — role-aware tab container
# ═══════════════════════════════════════════════════════════════════════════════

class MainWindow(QMainWindow):
    """
    Top-level window created after login.

    Roles
    -----
    admin    → both tabs: Document Q&A + Knowledge Hub
    engineer → Knowledge Hub only (no tab bar overhead)
    """

    def __init__(self, role: str, username: str = ""):
        super().__init__()
        self.role = role
        display_name = username.capitalize()
        role_label   = role.capitalize()
        self.setWindowTitle(f"AI Design Validator  —  {display_name}  ({role_label})")
        self.setMinimumSize(950, 680)
        self.setStyleSheet(self._get_stylesheet())

        if role == "admin":
            tabs = QTabWidget()
            tabs.setObjectName("MainTabWidget")
            tabs.addTab(DocuIntelligenceTab(),      "  📄  Docu Intelligence  ")
            tabs.addTab(DocuAssistanceTab(),  "  🧠  Docu Assistant  ")
            self.setCentralWidget(tabs)
        else:
            # Engineer sees Knowledge Hub directly — no tab chrome
            tabs = QTabWidget()
            tabs.setObjectName("MainTabWidget")
            tabs.addTab(DocuAssistanceTab(),  "  🧠  Docu Assistant  ")
            self.setCentralWidget(tabs)

    # ── Application-wide QSS ────────────────────────────────────────────────

    def _get_stylesheet(self) -> str:
        return """
        QMainWindow { background-color: #0f0f11; }

        /* ── Tab widget ── */
        QTabWidget#MainTabWidget::pane {
            border: none;
            background-color: #0f0f11;
        }
        QTabBar::tab {
            background-color: #16161a;
            color: #7f849c;
            border: 1px solid #232329;
            border-bottom: none;
            padding: 9px 22px;
            font-size: 12px;
            font-weight: 500;
        }
        QTabBar::tab:selected {
            background-color: #0f0f11;
            color: #89b4fa;
            border-top: 2px solid #7c4dff;
        }
        QTabBar::tab:hover:!selected { background-color: #1e1e2e; color: #cdd6f4; }

        /* ── Sidebar ── */
        QWidget#Sidebar {
            background-color: #16161a;
            border-right: 1px solid #232329;
        }
        QFrame#InstructionCard {
            background-color: #1e1e2e;
            border: 1px solid #313244;
            border-radius: 8px;
        }
        QLabel#BrandLabel {
            font-size: 20px;
            font-weight: bold;
            color: #89b4fa;
            margin-bottom: 6px;
        }
        QLabel#SectionTitle {
            font-size: 9px;
            font-weight: bold;
            color: #7f849c;
            margin-top: 14px;
            letter-spacing: 1px;
        }

        /* ── General controls ── */
        QPushButton {
            background-color: #2a2b36;
            color: #cdd6f4;
            border: 1px solid #3b3d52;
            border-radius: 6px;
            padding: 6px 12px;
            font-size: 12px;
            font-weight: 500;
        }
        QPushButton:hover   { background-color: #3b3d52; border-color: #585b70; }
        QPushButton:pressed { background-color: #1e1e2e; }

        QPushButton#UploadButton {
            background-color: #7c4dff;
            color: #fff;
            border: none;
            font-weight: bold;
            padding: 8px 12px;
        }
        QPushButton#UploadButton:hover   { background-color: #9d7cff; }
        QPushButton#UploadButton:pressed { background-color: #5d2bf0; }

        QPushButton#SecondaryButton {
            background-color: #181825;
            color: #bac2de;
            border: 1px solid #313244;
            text-align: left;
            padding: 6px 10px;
        }
        QPushButton#SecondaryButton:hover { background-color: #242437; border-color: #45475a; }

        /* ── File list ── */
        QScrollArea#FilesScrollArea { background-color: #11111b; border: 1px solid #232329; border-radius: 6px; }
        QWidget#FilesContainer      { background-color: #11111b; }
        QFrame#FileItemWidget       { background-color: #1e1e2e; border: 1px solid #313244; border-radius: 6px; }
        QFrame#FileItemWidget:hover { border-color: #45475a; }
        QPushButton#DeleteFileButton {
            background-color: transparent;
            color: #f38ba8;
            border: none;
            font-weight: bold;
        }
        QPushButton#DeleteFileButton:hover { background-color: #f38ba8; color: #11111b; border-radius: 3px; }

        /* ── Chat panel ── */
        QWidget#ChatPanel   { background-color: #0f0f11; }
        QFrame#ChatHeader   { background-color: #131317; border-bottom: 1px solid #232329; }
        QScrollArea#ChatScrollArea { background-color: #0f0f11; }
        QWidget#ChatContent        { background-color: #0f0f11; }

        QFrame#UserBubble {
            background-color: #2a2a35;
            border: 1px solid #3c3d4f;
            border-radius: 14px;
        }
        QFrame#AIBubble {
            background-color: #181822;
            border: 1px solid #23232e;
            border-radius: 14px;
        }

        QPushButton#BubbleExportButton {
            background-color: #242530;
            color: #a6adc8;
            border: 1px solid #313244;
            border-radius: 4px;
            font-size: 10px;
            padding: 2px 8px;
        }
        QPushButton#BubbleExportButton:hover { background-color: #89b4fa; color: #11111b; border-color: #89b4fa; }

        /* ── Input panel ── */
        QFrame#InputPanel { background-color: #131317; border-top: 1px solid #232329; }
        QTextEdit#MessageInput {
            background-color: #1c1c24;
            color: #eceff4;
            border: 1px solid #2e2e3a;
            border-radius: 8px;
            padding: 6px 12px;
            font-size: 13px;
        }
        QTextEdit#MessageInput:focus { border-color: #7c4dff; }
        QPushButton#AttachButton {
            background-color: #242530;
            color: #bac2de;
            border: 1px solid #313244;
            border-radius: 8px;
            font-size: 16px;
        }
        QPushButton#AttachButton:hover { background-color: #2e303f; }
        QPushButton#SendButton {
            background-color: #7c4dff;
            color: #fff;
            border: none;
            border-radius: 8px;
            font-weight: bold;
            font-size: 13px;
        }
        QPushButton#SendButton:hover   { background-color: #9d7cff; }
        QPushButton#SendButton:pressed { background-color: #5d2bf0; }

        /* ── Scrollbars ── */
        QScrollBar:vertical   { border:none; background:#11111b; width:8px; }
        QScrollBar::handle:vertical { background:#313244; min-height:20px; border-radius:4px; }
        QScrollBar::handle:vertical:hover { background:#45475a; }
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height:0; }
        QScrollBar:horizontal { border:none; background:#11111b; height:8px; }
        QScrollBar::handle:horizontal { background:#313244; min-width:20px; border-radius:4px; }
        QScrollBar::handle:horizontal:hover { background:#45475a; }
        QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width:0; }
        """


# ═══════════════════════════════════════════════════════════════════════════════
#  Entry point
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    from login import LoginWindow
    login = LoginWindow()
    login.show()
    sys.exit(app.exec())
